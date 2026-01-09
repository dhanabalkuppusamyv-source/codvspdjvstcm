import streamlit as st
import pandas as pd
import numpy as np
import os, io, re, unicodedata
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side


# ==============================
# Style & helpers (UI)
# ==============================
st.set_page_config(page_title="COD Compare", layout="wide")

logo_url = "https://raw.githubusercontent.com/Uthraa-18/cod-compare-app/refs/heads/main/image.png"

st.markdown(f"""
<style>
.corner-logo {{
    position: fixed;
    top: 50px;
    right: 20px;
    width: 120px;
    z-index: 9999;
}}
</style>

<img src="{logo_url}" class="corner-logo">
""", unsafe_allow_html=True)

st.markdown("""
<style>
.section-title {
    font-size: 1.35rem;
    font-weight: 700;
    display: inline-flex;
    align-items: center;
    gap: .5rem;
    margin: .25rem 0 .5rem 0;
}
.info-dot {
    display:inline-block;
    font-size: 0.95rem;
    line-height: 1;
    padding: .1rem .35rem;
    border-radius: 999px;
    border: 1px solid #aaa;
    color: #333;
    cursor: help;
}
.subtle {
    font-size: 0.95rem;
    color: #555;
    margin-top: .25rem;
}
.small-input .stNumberInput > div > div > input {
    font-size: .9rem;
}
.block-container { padding-top: 1rem; }
</style>
""", unsafe_allow_html=True)

def header_with_tip(text: str, tip: str):
    st.markdown(
        f"<div class='section-title'>{text}"
        f"<span class='info-dot' title='{tip}'>ⓘ</span></div>",
        unsafe_allow_html=True
    )

def extract_actual_nominal(nums, cod_nominal, eps):
    for x in nums:
        if approx_equal(x, cod_nominal, eps):
            return x
    return None


def extract_actual_tolerance(nums):
    """
    Return tolerance if present:
    - Prefer ± pair
    - Else single positive value
    """
    for x in nums:
        if -x in nums:
            return abs(x)
    for x in nums:
        if x > 0:
            return abs(x)
    return None

def extract_pdj_nominal(nums, cod_nominal, eps):
    """Return PDJ nominal value matching COD nominal"""
    for x in nums:
        if approx_equal(x, cod_nominal, eps):
            return x
    return ""

def extract_pdj_tolerance(nums):
    """Return PDJ tolerance (± value) if present"""
    for x in nums:
        if -x in nums:
            return fmt_pm(abs(x))
    return ""


def extract_all_images_from_cod(cod_file_path):
    wb = load_workbook(cod_file_path)
    ws = wb.active

    images = []

    if not hasattr(ws, "_images"):
        return images

    for i, img in enumerate(ws._images):
        img_path = f"/tmp/ref_image_{i}.png"

        # 🔥 Correct way: write raw image bytes
        with open(img_path, "wb") as f:
            f.write(img._data())

        images.append(img_path)

    return images


def first_numeric_in_row(df, r):
    nums = row_numbers(df, r)
    return nums[0] if nums else None



# ==============================
# Regex & utilities
# ==============================
RE_PM = re.compile(r'(?:±|\+/-)\s*(\d+(?:[.,]\d+)?)', re.I)
RE_SIGNED = re.compile(r'^[\+\-]?\s*\d+(?:[.,]\d+)?$')
RE_NUM = re.compile(r'[-+]?\d+(?:[.,]\d+)?')

def to_float(x):
    try:
        return float(str(x).replace(",", "."))
    except:
        return None

def norm(s):
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    s = str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    return s.lower().strip().replace("’","'")

def get_ext(name):
    return os.path.splitext(name)[-1].lower()

def read_all_sheets(name, file_bytes):
    engine = "xlrd" if get_ext(name)==".xls" else "openpyxl"
    xls = pd.ExcelFile(io.BytesIO(file_bytes), engine=engine)
    return {
        s: pd.read_excel(
            io.BytesIO(file_bytes),
            sheet_name=s,
            engine=engine,
            header=None
        )
        for s in xls.sheet_names
    }

# ==============================
# COD extraction helpers
# ==============================
def find_codification_value_below(cod_sheets, label="codification", scan_down=30):
    target = norm(label)
    for sname, df in cod_sheets.items():
        R, C = df.shape
        for r in range(R):
            for c in range(C):
                if norm(df.iat[r,c]) == target:
                    for rr in range(r+1, min(R, r+1+scan_down)):
                        if norm(df.iat[rr,c]) != "":
                            return sname, str(df.iat[rr,c]).strip(), r, c
    return None, None, None, None

def find_stacked_anchor_vertical(df, words, max_gap=10):
    R, C = df.shape
    W = [w.lower() for w in words]
    for c in range(C):
        starts = [r for r in range(R) if W[0] in norm(df.iat[r,c])]
        for r0 in starts:
            rcur = r0
            ok = True
            for w in W[1:]:
                found = False
                for rr in range(rcur+1, min(R, rcur+1+max_gap)):
                    if w in norm(df.iat[rr,c]):
                        rcur = rr
                        found = True
                        break
                if not found:
                    ok = False
                    break
            if ok:
                return rcur, c
    return None, None

def first_number_below(df, start_row, col, right_span=12, down_rows=4):
    R, C = df.shape
    for rr in range(start_row+1, min(R, start_row+1+down_rows)):
        for cc in range(col, min(C, col+right_span)):
            s = "" if pd.isna(df.iat[rr,cc]) else str(df.iat[rr,cc])

            m = RE_PM.search(s)
            if m:
                v = to_float(m.group(1))
                if v is not None:
                    return v, rr, cc

            if norm(s) in {"±","+/-"}:
                for cc2 in range(cc+1, min(C, cc+4)):
                    s2 = "" if pd.isna(df.iat[rr,cc2]) else str(df.iat[rr,cc2])
                    m2 = RE_NUM.search(s2)
                    if m2:
                        v = to_float(m2.group(0))
                        if v is not None:
                            return v, rr, cc2

            m3 = RE_NUM.search(s)
            if m3:
                v = to_float(m3.group(0))
                if v is not None:
                    return v, rr, cc
    return None, None, None

def two_signed_values_below_same_column(df, start_row, col, max_rows=8):
    R,_ = df.shape
    vals=[]
    for rr in range(start_row+1, min(R, start_row+1+max_rows)):
        s = "" if pd.isna(df.iat[rr,col]) else str(df.iat[rr,col]).strip()

        if RE_SIGNED.match(s):
            x = to_float(s)
            if x is not None:
                vals.append(x)
        elif norm(s) in {"±","+/-"} and col+1 < df.shape[1]:
            s2 = "" if pd.isna(df.iat[rr,col+1]) else str(df.iat[rr,col+1]).strip()
            if RE_NUM.match(s2):
                x = to_float(s2)
                if x:
                    vals.append(+abs(x))
                    vals.append(-abs(x))

    for v in vals:
        if -v in vals:
            return +abs(v), -abs(v)
    return None, None

# ==============================
# Extract numbers from PDJ/TCM rows
# ==============================
def row_numbers(df, r):
    nums=[]
    row = df.iloc[r,:].tolist()

    for i,v in enumerate(row):
        s = "" if pd.isna(v) else str(v)
        for m in RE_PM.findall(s):
            x = to_float(m)
            if x:
                nums += [+abs(x), -abs(x)]
        if norm(s) in {"±","+/-"}:
            for j in range(i+1, min(len(row), i+4)):
                s2 = "" if pd.isna(row[j]) else str(row[j])
                for m2 in RE_NUM.findall(s2):
                    x = to_float(m2)
                    if x:
                        nums += [+abs(x), -abs(x)]

    for v in row:
        s = "" if pd.isna(v) else str(v)
        for m in RE_NUM.findall(s):
            x = to_float(m)
            if x is not None:
                nums.append(x)

    return nums

def sheet_numbers(df):
    nums=[]
    for rr in range(df.shape[0]):
        nums += row_numbers(df, rr)
    return nums

def find_key_positions(df, key):
    key = str(key).strip()
    pos=[]
    R,C = df.shape
    for r in range(R):
        for c in range(C):
            s = "" if pd.isna(df.iat[r,c]) else str(df.iat[r,c]).strip()
            if s == key:
                pos.append((r,c))
    return pos

# ==============================
# Matching helpers
# ==============================
def approx_equal(a,b,tol):
    return abs(a-b) <= tol

def contains_value_eps(nums, val, tol):
    return any( approx_equal(x,val,tol) for x in nums )

def contains_pm_pair_eps(nums, mag, tol):
    return (
        any( approx_equal(x,+abs(mag),tol) for x in nums ) and
        any( approx_equal(x,-abs(mag),tol) for x in nums )
    )

def fmt_pm(m):
    s = f"{abs(m):.2f}".rstrip("0").rstrip(".")
    return f"+/- {s}"

# ==============================
# App UI
# ==============================
st.title("🔎 COD, PDJ, TCM Automatic Validation")

header_with_tip(
    "What this does",
    "Extracts Nominal & Tolerance from COD and compares PDJ/TCM rows."
)
st.caption("Epsilon allows 1.41 ≈ 1.4")

with st.container():
    st.markdown("<div class='small-input'>", unsafe_allow_html=True)
    eps = st.number_input(
        "Numeric tolerance (epsilon)",
        0.0,0.2,0.02,0.01,
        help="Lets ±1.41 match ±1.4"
    )
    st.markdown("</div>", unsafe_allow_html=True)

header_with_tip(
    "Upload COD workbook (.xls/.xlsx)",
    "Reads Codification, Nominal & Tolerance."
)
cod_file = st.file_uploader("", type=["xls","xlsx"], key="cod")

header_with_tip(
    "Upload PDJ/TCM/others",
    "PDJ + TCM → Only check row of key."
)
other_files = st.file_uploader(
    "",
    type=["xls","xlsx"],
    accept_multiple_files=True,
    key="others"
)

# ==============================
# Main logic
# ==============================
if cod_file and other_files:
    cod_bytes = cod_file.read()

    cod_temp_path = "/tmp/cod_source.xlsx"
    with open(cod_temp_path, "wb") as f:
        f.write(cod_bytes)

    ref_image_paths = extract_all_images_from_cod(cod_temp_path)

    
    cod_sheets = read_all_sheets(cod_file.name, cod_bytes)

    s_cod, key_value, _, _ = find_codification_value_below(cod_sheets,"codification")
    if not key_value:
        st.error("Could not find Codification value.")
        st.stop()

    st.markdown(
        f"<div class='subtle'>🔑 Compared Key: <code>{key_value}</code></div>",
        unsafe_allow_html=True
    )

    df_cod = cod_sheets[s_cod]

    nr, nc = find_stacked_anchor_vertical(df_cod, ["objectif","nominal","jeu"])
    if nr is None:
        st.error("Cannot find Objectif → Nominal → Jeu")
        st.stop()

    cod_nominal, _, _ = first_number_below(df_cod, nr, nc)
    if cod_nominal is None:
        st.error("Could not extract Nominal.")
        st.stop()

    tr, tc = find_stacked_anchor_vertical(df_cod, ["calcul","disp"])
    if tr is None:
        st.error("Cannot find Calcul → Disp.")
        st.stop()

    posv, negv = two_signed_values_below_same_column(df_cod, tr, tc)
    if posv is None or negv is None:
        pm, _, _ = first_number_below(df_cod, tr, tc)
        if pm is None:
            st.error("Could not extract Tolerance.")
            st.stop()
        tol_mag = abs(pm)
    else:
        tol_mag = abs(posv)

    ref_nom_disp = float(f"{cod_nominal:.2f}")
    ref_tol_disp = float(f"{tol_mag:.2f}")

    st.write(f"**COD Nominal (COD):** {ref_nom_disp}")
    st.write(f"**COD Tolerance (COD):** {fmt_pm(ref_tol_disp)}")

    # ============================
    # 4) Compare with other files
    # ============================
    results = []

    for f in other_files:
        f_bytes = f.read()
        f.seek(0)
        sheets = read_all_sheets(f.name, f_bytes)

        tag = f.name
        is_pdj = tag.upper().startswith("PDJ")
        is_tcm = tag.upper().startswith("TCM")

        for sname, df in sheets.items():

            pos = find_key_positions(df, key_value)
            if not pos:
                continue

            for (r, _) in pos:

                if is_pdj:
                    validation_nums = row_numbers(df, r)
                    tcm_row_nums = []

                # ---------- TCM ----------
                elif is_tcm:
                    validation_nums = row_numbers(df, r)
                    tcm_row_nums = validation_nums
                else:
                    row_nums = row_numbers(df, r)
                    if (
                        contains_value_eps(row_nums, cod_nominal, eps) or
                        contains_pm_pair_eps(row_nums, tol_mag, eps)
                    ):
                        validation_nums = row_nums
                    else:
                        validation_nums = sheet_numbers(df)
                    tcm_row_nums = []

                nominal_ok = contains_value_eps(validation_nums, cod_nominal, eps)
                tol_ok = contains_pm_pair_eps(validation_nums, tol_mag, eps)

                actual_nominal_found = extract_actual_nominal(validation_nums, cod_nominal, eps)
                actual_tolerance_found = extract_actual_tolerance(validation_nums)


                matched=[]
                if nominal_ok:
                    matched.append(f"{ref_nom_disp}")
                if tol_ok:
                    matched.append(fmt_pm(ref_tol_disp))
                pdj_nominal_val = ""
                pdj_tolerance_val = ""

                if is_pdj:
                    pdj_nominal_val = extract_pdj_nominal(validation_nums, cod_nominal, eps)
                    pdj_tolerance_val = extract_pdj_tolerance(validation_nums)


                results.append({
                    "Compared Key": key_value,
                    "File": tag,
                    "Sheet": sname,
                    "Key Row": r+1,
                    "COD Nominal": ref_nom_disp,
                    "COD Tolerance": fmt_pm(ref_tol_disp),
                    "PDJ Nominal Value": pdj_nominal_val,
                    "PDJ Tolerance Value": pdj_tolerance_val,
                    "TCM Nominal Value":
                        tcm_row_nums[0] if is_tcm and tcm_row_nums else "",
                    "TCM Tolerance Value":
                        fmt_pm(extract_actual_tolerance(tcm_row_nums))
                        if is_tcm and extract_actual_tolerance(tcm_row_nums) is not None
                        else "",

                    "Actual Nominal Found ?": "Yes" if nominal_ok else "No",
                    "Actual Tolerance Found ?": "Yes" if tol_ok else "No",
                    "OK - Nominal and Tolerance value": ", ".join(matched),
                    "Not-OK Value":
                        fmt_pm(actual_tolerance_found)
                        if actual_tolerance_found is not None and not tol_ok else "",
                })

    # ============================
    # 5) Final Table
    # ============================
    if results:

        df_out = pd.DataFrame(results)
        df_out["Ref image"] = ""
        # Add SI.No as first column
        df_out.insert(0, "SI.No", range(1, len(df_out) + 1))
        
        def color_yes_no(val):
            if val == "Yes":
                return "background-color:#C6F7C6; color:black;"
            else:
                return "background-color:#FFB3B3; color:black;"

        styled_df = df_out.style.applymap(
            color_yes_no,
            subset=["Actual Nominal Found ?", "Actual Tolerance Found ?"]
        )

        st.write("### 📊 Results")
        st.dataframe(styled_df, use_container_width=True)

        def create_colored_excel(df):
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Results"
            # Border
            thin = Side(style="thin")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            # Header
            ws.append(df.columns.tolist())

            for c in range(1, len(df.columns) + 1):
                ws.cell(row=1, column=c).border = border


            ref_col_idx = df.columns.get_loc("Ref image") + 1
            first_data_row = 2

            green = PatternFill(
                start_color="C6F7C6",
                end_color="C6F7C6",
                fill_type="solid"
            )
            red = PatternFill(
                start_color="FFB3B3",
                end_color="FFB3B3",
                fill_type="solid"
            )

            for row_idx, row_data in df.iterrows():
                ws.append([row_data[col] for col in df.columns])
                excel_r = ws.max_row

                for c in range(1, len(df.columns) + 1):
                    ws.cell(row=excel_r, column=c).border = border

                ws.cell(
                    excel_r,
                    df.columns.get_loc("Actual Nominal Found ?")+1
                ).fill = green if row_data["Actual Nominal Found ?"] == "Yes" else red

                ws.cell(
                    excel_r,
                    df.columns.get_loc("Actual Tolerance Found ?")+1
                ).fill = green if row_data["Actual Tolerance Found ?"] == "Yes" else red

            if ref_image_paths:
                col_letter = ws.cell(1, ref_col_idx).column_letter
                start_row = first_data_row

                for img_path in ref_image_paths:
                    img = XLImage(img_path)
                    img.width = 140
                    img.height = 90
                    ws.add_image(img, f"{col_letter}{start_row}")
                    ws.row_dimensions[start_row].height = 80
                    start_row += 1  # stack images vertically

                ws.column_dimensions[col_letter].width = 25

            output = io.BytesIO()
            wb.save(output)
            return output.getvalue()

        excel_data = create_colored_excel(df_out)

        st.download_button(
            "⬇️ Download Excel",
            excel_data,
            "cod_comparison_results.xlsx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    else:
        st.warning("No matches found.")
